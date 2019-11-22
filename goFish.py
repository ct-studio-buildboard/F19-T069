import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook

wb = xlrd.open_workbook('./team69.xlsx')  # Work Book
test = wb.sheet_by_name('Job Postings Dict')  # Work Sheet
positions = []
skills = []
for i in range(test.nrows):
	positions.append(test.cell(i,0))
	skills.append(test.cell(i,1))
jobs = {} #jobs dictionary
for j in range(len(skills)):
	#make dictionary add job position as key look at skills column and add that as value
	skill = str(skills[j])
	skill = skill.split(', ')
	jobs[positions[j]] = skill
print(jobs)

